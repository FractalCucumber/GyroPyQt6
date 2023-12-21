import logging
# from PyQt6.QtWidgets import QFileDialog
# from PyQt6.QtCore import pyqtSignal, QThread, QIODevice
# from pyqtgraph.Qt import QtCore, QtGui
# from datetime import datetime
# from PyQt6 import QtWidgets, QtCore, QtGui
from PyQt5 import QtWidgets, QtCore, QtGui
import pyqtgraph.exporters
import pyqtgraph as pg
import numpy as np
from openpyxl import load_workbook  # лучше использовать Pandas, наверняка работает быстрее
import os
import sys
# pg.setConfigOption('background', '#f0f0f5')  # Установите фон в серый цвет
# pg.setConfigOption('foreground', 'd')

# QtWidgets.QDialog.create
class CustomDialog(QtWidgets.QDialog):
    # close_emit = QtCore.pyqtSignal(bool)

    def __init__(self):
        super().__init__()
        STYLE_SHEETS_FILENAME = 'res\StyleSheets2.css'
        with open(get_res_path(STYLE_SHEETS_FILENAME), "r") as style_sheets:
            self.setStyleSheet(style_sheets.read())
        self.setMaximumSize(500, 175)
        self.setWindowTitle("Окно изменения проектов")
        self.setWindowFlags(QtCore.Qt.WindowType.CustomizeWindowHint |
                            QtCore.Qt.WindowType.WindowCloseButtonHint)
        self.QBtn = QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel
        self.button_box = QtWidgets.QDialogButtonBox(self.QBtn)
        self.button_box.button(
            QtWidgets.QDialogButtonBox.Ok).setEnabled(False)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)

        self.layout = QtWidgets.QGridLayout()
        message = QtWidgets.QLabel("Выберите путь к файлу и назовите проект",
                                   wordWrap=True, maximumHeight=80)
        self.layout.addWidget(message, 0, 0, 1, 3)
        name_label = QtWidgets.QLabel("Имя")
        self.layout.addWidget(name_label, 1, 0, 1, 1)
        self.name = QtWidgets.QLineEdit()
        self.name.textChanged.connect(self.check)
        self.layout.addWidget(self.name, 1, 1, 1, 2)
        path_label = QtWidgets.QLabel("Путь")
        self.layout.addWidget(path_label, 2, 0, 1, 1)
        self.path = QtWidgets.QLineEdit()
        self.layout.addWidget(self.path, 2, 1, 1, 1)
        self.path.textChanged.connect(self.check)
        open_folder_btn = QtWidgets.QPushButton()
        btn_icon = QtGui.QIcon(get_res_path('res//open_folder.png'))
        # btn_icon.addFile(self.get_res_path('res//open_folder.png'), QtCore.QSize(48, 48))
        open_folder_btn.setIcon(btn_icon)
        self.layout.addWidget(open_folder_btn, 2, 2, 1, 1)
        open_folder_btn.clicked.connect(self.get_path)
        self.layout.addWidget(self.button_box, 3, 0, 1, 3)
        self.setLayout(self.layout)

    def check(self):
        if os.path.exists(self.path.text()) and len(self.name.text()):
            self.button_box.button(
                QtWidgets.QDialogButtonBox.Ok).setEnabled(True)
        else:
            self.button_box.button(
                QtWidgets.QDialogButtonBox.Ok).setEnabled(False)
    # def closeEvent(self, a0):
        # if [self.name.text(), str(self.filename)]:
            # self.close_emit.emit(True)
        # [self.name.text(), str(self.filename)]
        # return super().closeEvent(a0)

    def get_path(self):
        self.filename, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Выберите файл",
            ".",
            "Excel Files (*.xls *.xlsx)")
        self.path.setText(str(self.filename))

# @staticmethod
def get_res_path(relative_path):
    """
    Get absolute path to resource, works with PyInstaller
    """
    base_path = getattr(
        sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

class CustomViewBox(pg.ViewBox):
    def __init__(self, *args, **kwds):
        kwds['enableMenu'] = True
        # kwds['enableMenu'] = False
        pg.ViewBox.__init__(self, *args, **kwds)
        # self.setMouseMode(self.RectMode)

    #  reimplement right-click to zoom out
    def mouseClickEvent(self, ev):
        if ev.button() == QtCore.Qt.MouseButton.RightButton:
            self.autoRange()

    # #  reimplement mouseDragEvent to disable continuous axis zoom
    # def mouseDragEvent(self, ev, axis=None):
    #     if axis is not None and ev.button() == QtCore.Qt.MouseButton.RightButton:
    #         ev.ignore()
    #     else:
    #         pg.ViewBox.mouseDragEvent(self, ev, axis=axis)


class CustomTabWidget(QtWidgets.QTabWidget):
    warning_signal = QtCore.pyqtSignal(str)
    get_filename_signal = QtCore.pyqtSignal(bool)
    # filenames_list_emit = QtCore.pyqtSignal(bool)

    def __init__(self, fs=1000, GYRO_NUMBER=1, logger_name='', parent=None):
        # QtWidgets.QTabWidget.__init__(self)
        super(CustomTabWidget, self).__init__(parent)
        # QtWidgets.QWidget.__init__(self, parent)
        self.GYRO_NUMBER = GYRO_NUMBER
        self.visibility_flags = [True for _ in range(self.GYRO_NUMBER + 1)]
        self.LABEL_STYLE = {'color': '#FFF', 'font-size': '16px'}
        self.COLOR_LIST = ['r', 'g', 'b']
        self.fs = fs
        self.dlg = CustomDialog()
        self.filenames_to_fft = []
        self.dict = {}
        # self.settings = QtCore.QSettings("settings")
# ------ time plot ------------------------------------------------------------

        self.time_plot_item = pg.PlotItem(viewBox=CustomViewBox())
        self.time_plot = pg.PlotWidget(plotItem=self.time_plot_item)
        self.time_plot_item.setTitle('Угловая скорость', size='13pt')  # Velosity Graph
        self.time_plot_item.showGrid(x=True, y=True)
        self.time_plot_item.addLegend(offset=(-1, 1), labelTextSize='12pt',
                                      labelTextColor=pg.mkColor('w'))
        self.time_plot_item.setLabel('left', 'Velosity',
                                     units='\u00b0/second', **self.LABEL_STYLE)
        self.time_plot_item.setLabel('bottom', 'Time',
                                     units='seconds', **self.LABEL_STYLE)

        self.time_curves = [self.time_plot_item.plot(pen='w', name="encoder")]
        for i in range(self.GYRO_NUMBER):
            self.time_curves.append(self.time_plot_item.plot(
                pen=self.COLOR_LIST[i], name=f"gyro {i + 1}"))

        self.region = pg.LinearRegionItem([0, 1], movable=False)
        self.time_plot_item.addItem(self.region)

        # x = [1, 5, 10, 15, 20, 25, 33, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 100, 105, 110, 115, 118, 120, 122, 125, 130, 135, 140, 150, 156, 162, 170, 180, 190, 200, 205, 210, 215, 220, 225, 230, 235, 240, 245, 250, 255, 260, 265, 270, 275, 280, 285, 290, 300, 310]
        # self.x = [1, 5, 10, 15, 20, 25, 33, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 100, 102, 105, 107, 110, 115, 118, 120, 122, 125, 130, 135, 140, 150, 156, 162, 170, 180, 190, 200, 205, 210, 215, 220, 225, 230, 235, 240, 243, 247, 250, 255, 260, 265, 270, 275, 280, 285, 290, 295, 300, 305, 310, 315]
        # # # amp =np.array([96, 97, 108, 112, 113, 114, 121, 127, 136, 163, 166, 191, 219, 240, 258, 284, 400, 371, 450, 699, 791, 1154, 631, 697, 722, 743, 834, 823, 918, 995, 1022, 1125, 1220, 1244, 1373, 1468, 1618, 1851, 1865, 2166, 2548, 2539, 3409, 3521, 3514, 4220, 3473, 2573, 3081, 3028, 3028, 3230, 3056, 3132, 3102, 3621, 3669, 4561])/100
        # self.y =np.array(
        #     [94, 99, 105, 114, 112, 113, 122, 127, 137, 160, 169, 205, 221, 243, 292, 306, 339, 392, 419, 554, 861, 913, 1079, 1276, 595, 645, 659, 666, 781, 810, 863, 948, 1024, 1103, 1227, 1329, 1319, 1362, 1468, 1791, 1921, 1959, 2641, 2477, 3384, 3484, 3254, 4482, 3888, 3931, 2795, 3021, 3246, 3046, 3147, 3505, 3424, 3430, 4049, 4157, 4385, 3980, 5340, 4897]
        #     )/100
        # self.time_curves[0].setData(self.x, self.y)
# ------ Tab widget -----------------------------------------------------------
        self.page = QtWidgets.QWidget(self)
        self.layout = QtWidgets.QVBoxLayout()
        self.layout.addWidget(self.time_plot)
        self.spectrum_button = QtWidgets.QPushButton("От времени")  # Time plot
        self.layout.addWidget(self.spectrum_button)
        self.page.setLayout(self.layout)
        self.addTab(self.page, "\u03C9(t)")  # &Time plot От времени
        self.phase_curves: list[pg.PlotCurveItem] = []
        self.amp_curves: list[pg.PlotCurveItem] = []
        self.phase_plot_list: list[pg.PlotWidget] = []
        self.amp_plot_list: list[pg.PlotWidget] = []
        self.tab_widget_page_list: list[QtWidgets.QWidget] = []

        self.logger = logging.getLogger(logger_name)
        self.spectrum_button.clicked.connect(self.switch_plot_x_axis)
# -----------------------------------------------------------------------------
        self.groupbox = QtWidgets.QGroupBox(
            ' Сохранение в .xlsx', maximumWidth=190, minimumWidth=150)
        self.median_plot_groupbox_layout = QtWidgets.QGroupBox(
            'Содержимое файла', maximumWidth=315, minimumWidth=215)
        self.median_plot_groupbox_layout = QtWidgets.QGridLayout(spacing=10)
        # self.median_plot_groupbox_layout.setRowStretch(4, 0)
        # self.median_plot_groupbox_layout.setSpacing(0)
        # self.median_plot_groupbox_layout.setContentsMargins(0, 0, 0, 0)
        self.groupbox.setLayout(self.median_plot_groupbox_layout)

        self.Q_label = QtWidgets.QLineEdit('Q=')
        self.f_r_label = QtWidgets.QLineEdit('f_r=')
        self.f_r_label = QtWidgets.QLineEdit()
        # self.layout.addWidget(self.Q_label)
        self.projects_combo_box = QtWidgets.QComboBox(editable=True)
        self.projects_combo_box.lineEdit().setReadOnly(True)
        # QToolButton # есть wordWrap

        # if self.dict:
        #     self.test_combo_box.addItems(self.dict.keys())
        #     for i in range(self.test_combo_box.count()):
        #         self.test_combo_box.setItemData(
        #             i, self.dict.get(self.test_combo_box.itemText(i)),
        #             QtCore.Qt.ItemDataRole.ToolTipRole)
        self.median_plot_groupbox_layout.addWidget(
            self.projects_combo_box, 0, 0, 1, 3)

        self.amp_info_label = QtWidgets.QLabel("amp")
        self.median_plot_groupbox_layout.addWidget(
            self.amp_info_label, 2, 0, 1, 1)
        self.amp_info = QtWidgets.QLineEdit("0")
        self.median_plot_groupbox_layout.addWidget(
            self.amp_info, 2, 1, 1, 2)

        self.freq_info_label = QtWidgets.QLabel("f, Hz")
        self.median_plot_groupbox_layout.addWidget(
            self.freq_info_label, 3, 0, 1, 1)
        self.freq_info = QtWidgets.QLineEdit("0")
        self.median_plot_groupbox_layout.addWidget(
            self.freq_info, 3, 1, 1, 2)

        self.name_info_label = QtWidgets.QLabel("имя датчика:",
                                                maximumHeight=40)
        self.median_plot_groupbox_layout.addWidget(
            self.name_info_label, 4, 0, 1, 3)
        self.name_info = QtWidgets.QLineEdit('folder')
        self.median_plot_groupbox_layout.addWidget(
            self.name_info, 5, 0, 1, 3)

        self.ok_btn = QtWidgets.QPushButton("Запись в Excel")
        self.median_plot_groupbox_layout.addWidget(self.ok_btn, 6, 0, 1, 3)
        self.ok_btn.clicked.connect(self.write_xlsx)
        # self.ok_btn.setSizePolicy(
        #     QtWidgets.QSizePolicy.Policy.MinimumExpanding,
        #     QtWidgets.QSizePolicy.Policy.Maximum)
        self.add_btn = QtWidgets.QPushButton(objectName='add')
        self.add_btn.setFixedSize(30, 30)
        btn_icon = QtGui.QIcon(get_res_path('res//add.png'))
        self.add_btn.setIcon(btn_icon)
        self.median_plot_groupbox_layout.addWidget(self.add_btn, 1, 0, 1, 1)
        self.add_btn.clicked.connect(self.change_xlsx_list)
        self.del_btn = QtWidgets.QPushButton(objectName="del")
        self.del_btn.setFixedSize(30, 30)
        btn_icon = QtGui.QIcon(get_res_path('res//delete2.png'))
        # btn_icon.addFile(, QtCore.QSize(48, 48))
        self.del_btn.setIcon(btn_icon)
        self.median_plot_groupbox_layout.addWidget(self.del_btn, 1, 1, 1, 1)
        self.del_btn.clicked.connect(self.change_xlsx_list)  # del_xlsx_item
        self.change_btn = QtWidgets.QPushButton(objectName="change")
        self.change_btn.setFixedSize(30, 30)
        # self.setIconSize(QtCore.QSize(48, 48))
        btn_icon = QtGui.QIcon(get_res_path('res//edit.png'))
        # btn_icon.addFile(, QtCore.QSize(128, 128))
        self.change_btn.setIcon(btn_icon)
        self.median_plot_groupbox_layout.addWidget(self.change_btn, 1, 2, 1, 1)
        self.change_btn.clicked.connect(self.change_xlsx_list)

        self.get_filenames_btn = QtWidgets.QPushButton("АФЧХ для файла")
        self.median_plot_groupbox_layout.addWidget(
            self.get_filenames_btn, 7, 0, 1, 3)
        self.get_filenames_btn.clicked.connect(self.get_filenames)

        self.plot_fft_median()
# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------
#
###############################################################################
#
# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------
# ----- plotting --------------------------------------------------------------

    def plot_time_graph(self, time_data: np.ndarray,
                        enc_data: np.ndarray, gyro_data: np.ndarray):
        self.time_curves[0].setData(time_data, enc_data)
        for i in range(self.GYRO_NUMBER):
            self.time_curves[i + 1].setData(time_data, gyro_data)

    def set_fft_data(self, freq_data: np.ndarray, frame: list, fs: int):
        """Adds points to frequency graphs"""
        self.logger.info("plot_fft")
        for i in range(self.GYRO_NUMBER):
            self.amp_curves[-1 - i].setData(freq_data[:, 0],
                                        freq_data[:, 1])
            self.phase_curves[-1 - i].setData(freq_data[:, 0],
                                        freq_data[:, 2])
        self.region.setRegion([frame[0]/fs, frame[1]/fs])
        self.amp_plot_list[-1].autoRange()

# ------ FFT median plot ------------------------------------------------------------------
    def set_fft_median_data(self, freq_data: np.ndarray,
                            special_points: np.ndarray, folder: str):
        self.logger.info(f"Final median plot, sensor: {folder}")
        self.freq_data = freq_data
        for i in range(self.GYRO_NUMBER):
            self.amp_curves[i].setData(freq_data[:, -4],
                                        freq_data[:, -3])
            self.phase_curves[i].setData(freq_data[:, -4],
                                        freq_data[:, -2])
        # self.add_180_2d_scatter(self. plot, special_points[0], special_points[1])
        self.amp_info.setText(
            str(round(np.max(freq_data[:, -3]), 3)))
        # self.amp_info.setText(special_points[1])
        self.freq_info.setText(
            str(round(freq_data[np.argmax(freq_data[:, -3]), -4], 3)))
        # self.freq_info.setText(special_points[2])
        self.name_info.setText(folder)
        self.amp_plot_list[0].autoRange()
        self.mouse_x = None
        self.mouse_y = None

    def plot_fft_median(self):
        self.freq_data = np.array([])
        self.last_tab_layout = QtWidgets.QGridLayout(spacing=0)
        self.last_tab_layout.addWidget(self.groupbox, 0, 1, 2, 1) 

        # index = self.count() - 1
        self.tab_widget_page_list.append(QtWidgets.QWidget(self))
        self.append_amp_plot()
        self.last_tab_layout.addWidget(self.amp_plot_list[0], 0, 0, 1, 1)
        self.append_phase_plot()
        self.last_tab_layout.addWidget(self.phase_plot_list[0], 1, 0, 1, 1)
        self.tab_widget_page_list[0].setLayout(self.last_tab_layout)
        self.addTab(
            self.tab_widget_page_list[0], "&АФЧХ (средний)")  # FC average
        # self.setCurrentIndex(1)
  
        # # amp =np.array([96, 97, 108, 112, 113, 114, 121, 127, 136, 163, 166, 191, 219, 240, 258, 284, 400, 371, 450, 699, 791, 1154, 631, 697, 722, 743, 834, 823, 918, 995, 1022, 1125, 1220, 1244, 1373, 1468, 1618, 1851, 1865, 2166, 2548, 2539, 3409, 3521, 3514, 4220, 3473, 2573, 3081, 3028, 3028, 3230, 3056, 3132, 3102, 3621, 3669, 4561])/100
        # self.plot_2d_scatter(self.time_plot, self.x, self.y)
        # self.cursor = QtCore.Qt.BlankCursor
        self.amp_plot_cursor = QtCore.Qt.CrossCursor # was
        self.amp_plot_list[0].setCursor(self.amp_plot_cursor)  # self.cursor ##########################
        self.phase_plot_cursor = QtCore.Qt.CrossCursor # was
        self.phase_plot_list[0].setCursor(self.phase_plot_cursor)  # self.cursor ##########################

        # Add lines
        self.infinite_line_x1 = pg.InfiniteLine(angle=90, movable=False)
        self.infinite_line_x2 = pg.InfiniteLine(angle=90, movable=False)
        # self.infinite_line_y = pg.InfiniteLine(angle=0, movable=False)
        self.amp_plot_list[0].addItem(self.infinite_line_x1, ignoreBounds=True)
        self.phase_plot_list[0].addItem(self.infinite_line_x2, ignoreBounds=True)
        # self.amp_plot_list[0].addItem(self.infinite_line_y, ignoreBounds=True)
        self.cursorlabel = pg.TextItem()
        self.cursorlabel.setPos(3, 0.8)

        self.amp_plot_list[0].addItem(self.cursorlabel) #################################
        self.proxy_amp = pg.SignalProxy(
            self.amp_plot_list[0].scene().sigMouseMoved, delay=0.1, ##############################
            rateLimit=12, slot=self.update_crosshair)
        self.proxy_phase = pg.SignalProxy(
            self.phase_plot_list[0].scene().sigMouseMoved, delay=0.1, ##############################
            rateLimit=12, slot=self.update_crosshair)
        self.amp_plot_list[0].autoRange()
        # self.proxy = pg.SignalProxy(
            # self.phase_plot_list[-1].scene().sigMouseMoved, delay=0.15,
            # rateLimit=15, slot=self.update_crosshair) ####################

    def add_180_2d_scatter(self, plot, x, y, color=(66, 245, 72)):
        brush = pg.mkBrush(color)
        scatter = pg.ScatterPlotItem(size=5, brush=brush)
        scatter.addPoints(x, y)
        # scatter.setData(x, y)  # ??
        plot.addItem(scatter)

# ----- cursor ---------------------------------------------------------------------------------
    def update_crosshair(self, e):
        pos = e[0]
        if (self.amp_plot_list[0].plotItem.sceneBoundingRect().contains(pos)) and self.freq_data.size:
            mousePoint = self.amp_plot_list[0].plotItem.vb.mapSceneToView(pos)
            # mx = np.array(
                # [abs(float(i) - float(mousePoint.x())) for i in self.x])
            # index = mx.argmin()
            # for i in range(self.GYRO_NUMBER):
            #     if self.visibility_flags[i]:
            #         # ищем совпадение для i-го гироскопа
            #         break
            index = np.argmin(np.abs(self.freq_data[:, -4] - mousePoint.x()))
            # if index >= 0 and index < len(self.x):
            if index >= 0 and index < len(self.freq_data[:, -4]):
                self.cursorlabel.setText(f"f {self.freq_data[index, -4]:.3f}\n" +
                                         f"A {self.freq_data[index, -3]:.2f}\n" +
                                         f"\u03C6 {self.freq_data[index, -2]:.1f}")
                # self.infinite_line_x.setPos(self.x[index])
                self.infinite_line_x1.setPos(self.freq_data[index, -4]) 
                self.infinite_line_x2.setPos(self.freq_data[index, -4]) 
                # self.mouse_x = self.infinite_line_x.setPos(self.x[index])
                # self.mouse_y = self.infinite_line_y.setPos(self.y[index])
                self.mouse_x = (self.freq_data[index, -4]) # freq_data[:, -4]
                self.mouse_y = (self.freq_data[index, -3]) # freq_data[:, -3]
                # self.scatter.addPoints(self.mouse_x, self.mouse_y)
                # self.cursorlabel.setText(
                #         str((self.x[index], self.y[index])))
                # self.infinite_line_x.setPos(self.x[index])
                # self.infinite_line_y.setPos(self.y[index]) 
                # self.mouse_x = self.infinite_line_x.setPos(self.x[index])
                # self.mouse_y = self.infinite_line_y.setPos(self.y[index])
                # self.mouse_x = (self.x[index]) # freq_data[:, -4]
                # self.mouse_y = (self.y[index]) # freq_data[:, -3]
            
    # def mousePressEvent(self, e):
    def mouseDoubleClickEvent(self, e):
        if (e.buttons() & QtCore.Qt.LeftButton & (
            self.amp_plot_list[0].underMouse() | self.phase_plot_list[0].underMouse()
            )) and self.freq_data.size: #| self.phase_plot_list[-1].underMouse()):
            self.amp_info.setText(f'{self.mouse_y:.3f}')
            self.freq_info.setText(f'{self.mouse_x:.2f}')
            # self.cursorlabel.setPos(self.mouse_x + 1, self.mouse_y + 1)
            # if self.mouse_x in self.x and self.mouse_y in self.y:

# ----- Excel and file selecting ----------------------------------------------
    def get_filenames(self):
        self.filenames_to_fft, _ = QtWidgets.QFileDialog.getOpenFileNames(
            None,
            "QFileDialog.getOpenFileNames()",
            "",
            "Text Files (*.txt)")
        if len(self.filenames_to_fft):
            self.get_filename_signal.emit(True)
            # "All Files (*);;Python Files (*.py);;Text Files (*.txt)")

    # def del_xlsx_item(self):
    #     self.dict.pop(self.projects_combo_box.currentText(), None)
    #     self.projects_combo_box.removeItem(self.projects_combo_box.currentIndex())

    def change_xlsx_list(self):  # можно и del_xlsx сюда впихнуть
        if self.sender().objectName() == 'del':
            self.dict.pop(self.projects_combo_box.currentText(), None)
            self.projects_combo_box.removeItem(self.projects_combo_box.currentIndex())
    
        if self.sender().objectName() == 'change' and not self.projects_combo_box.count():
            return
        if self.sender().objectName() == 'change':
            self.dlg.path.setText(self.dict[self.projects_combo_box.currentText()])
            self.dlg.name.setText(self.projects_combo_box.currentText())
        if self.dlg.exec():
            if self.sender().objectName() == 'add':
                self.projects_combo_box.insertItem(0, self.dlg.name.text())
                self.projects_combo_box.setCurrentIndex(0)
            if self.sender().objectName() == 'change':
                self.dict.pop(self.projects_combo_box.currentText(), None)
            self.dict[self.dlg.name.text()] = self.dlg.path.text()
            self.projects_combo_box.setCurrentText(self.dlg.name.text())
            self.projects_combo_box.setItemText(self.projects_combo_box.currentIndex(),
                                            self.dlg.name.text())
            self.projects_combo_box.setItemData(self.projects_combo_box.currentIndex(),
                                            self.dict[self.dlg.name.text()],
                                            QtCore.Qt.ItemDataRole.ToolTipRole)

    # def get_new_item(self, flag):
        # self.test_combo_box.setCurrentText(self.dlg.name.text())
        # self.dlg.filename
        # print('close!')
 
    def write_xlsx(self):
        self.currnt_xlsx_path = self.dict[self.projects_combo_box.currentText()]
        self.currnt_xlsx_path = "test.xlsx"
        try:
            wb = load_workbook(self.currnt_xlsx_path)
        except FileNotFoundError:
            self.warning_emit.emit("File not found!")
            return
        ws = wb['Лист1']  # or wb.active
        ws['F3'] = self.name_info.text()
        ws['F4'] = self.freq_info.text()
        ws['F5'] = self.amp_info.text()
        # ws.column_dimensions['F'].width = 51
        try:
            wb.save(self.currnt_xlsx_path)
        except IOError:
            self.warning_emit.emit("File was open! Close and try again")

# ----- plot change -----------------------------------------------------------

    def clear_plots(self):
        self.freq_data = np.array([])
        for i in range(self.count() - 2):
            self.removeTab(2)
        #     self.tab_widget_page_list.pop()
        #     for i in range(self.GYRO_NUMBER):
            self.amp_curves.pop()
            self.phase_plot_list.pop()
            self.amp_plot_list.pop()
            self.tab_widget_page_list.pop()
        #         phase_plot_list
        self.time_curves[0].setData([])
        for i in range(self.GYRO_NUMBER):
            self.time_curves[i + 1].setData([])
            self.amp_curves[i].setData([])
            self.phase_curves[i].setData([])
        # self.phase_curves: list[pg.PlotCurveItem] = []
        # self.amp_curves: list[pg.PlotCurveItem] = []
        # self.phase_plot_list: list[pg.PlotWidget] = []
        # self.amp_plot_list: list[pg.PlotWidget] = []
        # self.tab_widget_page_list: list[QtWidgets.QWidget] = []

    @QtCore.pyqtSlot()
    def switch_plot_x_axis(self):
        """
        Switching between time plot and spectrum
        """
        if self.spectrum_button.text() == "Спектр":  # Frequency plot
            self.spectrum_button.setText("От времени")  # Time plot
            self.time_plot_item.ctrl.fftCheck.setChecked(False)
            self.time_plot_item.setLabel(
                'bottom', 'Time', units='seconds')
            self.region.show()
        else:
            self.spectrum_button.setText("Спектр")  # Frequency plot
            self.time_plot_item.ctrl.fftCheck.setChecked(True)
            self.time_plot_item.setLabel(
                'bottom', 'Frequency', units='Hz')
            self.region.hide()

    def change_curve_visibility(self):
        ind = int(self.sender().objectName())
        # self.flag_visibility[num] = self.check_box_list[num].checkState()
        self.visibility_flags[ind] = not self.visibility_flags[ind]
        self.time_curves[ind].setVisible(self.visibility_flags[ind])
        if ind == 0:
            return
        for i in range(self.count() - 1):
            self.phase_curves[i - 1 + ind].setVisible(
                self.visibility_flags[ind])
            self.amp_curves[i - 1 + ind].setVisible(
                self.visibility_flags[ind])

    def append_fft_plot_tab(self):
        """
        Create new tab and append amplitude and grequency graphs
        """
        index = self.count() - 1
        self.tab_widget_page_list.append(QtWidgets.QWidget(self))
        self.layout = QtWidgets.QVBoxLayout(spacing=0)
        self.append_amp_plot()
        self.layout.addWidget(self.amp_plot_list[index])
        self.append_phase_plot()
        self.layout.addWidget(self.phase_plot_list[index])
        self.tab_widget_page_list[-1].setLayout(self.layout)
        self.addTab(
            self.tab_widget_page_list[-1], f"ЧХ &{index}")  # FC        

    def append_amp_plot(self):
        self.amp_plot_item = pg.PlotItem(viewBox=CustomViewBox(),
                                         name=f'amp_plot{self.count()}')
        self.amp_plot_item.setXLink(f'phase_plot{self.count()}')
        self.amp_plot_item.setTitle('АФЧХ', size='12pt')  # Amp Graph
        self.amp_plot_item.showGrid(x=True, y=True)
        self.amp_plot_item.addLegend(offset=(-1, 1), labelTextSize='12pt',
                                     labelTextColor=pg.mkColor('w'))
        self.amp_plot_item.setLabel('left', 'Amplitude',
                                    units="", **self.LABEL_STYLE)
        # self.amp_plot_item.setLabel('bottom', 'Frequency',
        #                             units='Hz', **self.LABEL_STYLE)
        # self.SYMBOL_SIZE = 6
        for i in range(self.GYRO_NUMBER):
            self.amp_curves.append(self.amp_plot_item.plot(
                pen=self.COLOR_LIST[i], name=f"gyro {i + 1}", symbol="o",
                symbolSize=6, symbolBrush=self.COLOR_LIST[i]))
        self.amp_plot_list.append(pg.PlotWidget(plotItem=self.amp_plot_item))
        self.amp_plot_list[-1].getAxis('left').setWidth(60)
        self.amp_plot_list[-1].setLimits(xMin=-5, xMax=int(self.fs * 0.53), yMin=-0.08, yMax=100)

    def append_phase_plot(self):
        self.phase_plot_item = pg.PlotItem(viewBox=CustomViewBox(),
                                           name=f'phase_plot{self.count()}')
        self.phase_plot_item.setXLink(f'amp_plot{self.count()}')
        # self.phase_plot_item.setTitle('ФЧХ', size='12pt')  # Phase Graph
        self.phase_plot_item.showGrid(x=True, y=True)
        self.phase_plot_item.addLegend(offset=(-1, 1), labelTextSize='12pt',
                                       labelTextColor=pg.mkColor('w'))
        self.phase_plot_item.setLabel('left', 'Phase',
                                      units='degrees', **self.LABEL_STYLE)  # rad
        self.phase_plot_item.setLabel('bottom', 'Frequency',
                                      units='Hz', **self.LABEL_STYLE) # \u00b0
        for i in range(self.GYRO_NUMBER):
            self.phase_curves.append(self.phase_plot_item.plot(
                pen=self.COLOR_LIST[i], name=f"gyro {i + 1}", symbol="o",
                symbolSize=6, symbolBrush=self.COLOR_LIST[i]))
        self.phase_plot_list.append(
            pg.PlotWidget(plotItem=self.phase_plot_item))
        self.phase_plot_list[-1].getAxis('left').setWidth(60)
        self.phase_plot_list[-1].setLimits(
            xMin=-5, xMax=int(self.fs * 0.53), yMin=-375, yMax=20)

    @QtCore.pyqtSlot(str)
    def save_plot_image(self, path):
        # self.logger.info("Save image")
        pyqtgraph.exporters.ImageExporter(
            self.time_plot_item).export(path + '_time_plot.png')
        for i in range(self.count() - 1):
            pyqtgraph.exporters.ImageExporter(
                self.amp_plot_list[i].getPlotItem()).export(
                    path + f'_amp_plot_{i + 1}.png')
            pyqtgraph.exporters.ImageExporter(
                self.phase_plot_list[i].getPlotItem()).export(
                    path + f'_phase_plot_{i + 1}.png')
        # self.logger.info("Saving complite")

# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
#
###############################################################################
#
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------


class CustomComboBox(QtWidgets.QComboBox):
    def __init__(self,
                 settings: QtCore.QSettings,  # передавать сразу объект QtCore.QSettings, чтобы не создавать их
                 settings_name: str = 'Deafault',
                 default_items_list = [''],
                 editable_flag=True,
                 uint_validator_enable=True,
                 parent=None):
        super(CustomComboBox, self).__init__(parent)
        self.settings_name = settings_name
        self.setEditable(True)
        self.lineEdit().setReadOnly(not editable_flag)
        self.lineEdit().setAlignment(
            QtCore.Qt.AlignmentFlag.AlignCenter)
        if uint_validator_enable:
            self.int_validator = QtGui.QIntValidator(bottom=0)
            self.setValidator(self.int_validator)

        self.currentTextChanged.connect(
            lambda value: self.setItemText(self.currentIndex(), value))

        # self.settings = QtCore.QSettings(settings_name)
        self.settings = settings
        if self.settings.contains("items" + self.settings_name):
            self.addItems(
                self.settings.value("items" + self.settings_name))
        else:
            if not len(default_items_list[0]):
                return
            
            self.addItems(default_items_list)
        if self.settings.contains("curr_index" + self.settings_name):
            self.setCurrentIndex(
                self.settings.value("curr_index" + self.settings_name))
        if self.settings.contains("name" + self.settings_name):
            for i in range(self.count()):
                if self.itemText(i) == self.settings.value("name" + self.settings_name):
                    self.setCurrentIndex(i)
                    break

        # self.save_value = lambda: self.settings.setValue(
        #     "items",
        #     [self.itemText(i) for i in range(self.count())])

        # self.save_index = lambda: self.settings.setValue(
        #     "curr_index", self.currentIndex())

        # self.save_name = lambda: self.settings.setValue(
        #     "COM_current_name", self.currentText())

    def save_all(self):
        self.save_value()
        self.save_index()

    def save_value(self):
        if self.count():
            self.settings.setValue(
                "items" + self.settings_name,
                [self.itemText(i) for i in range(self.count())])

    def save_index(self):
        if self.count():
            self.settings.setValue(
                "curr_index" + self.settings_name, self.currentIndex())

    def save_current_text(self):
        if self.count():
            self.settings.setValue(
                "name" + self.settings_name, self.currentText())

# ----------------------------------------------------------------------------
#
# ----------------------------------------------------------------------------
#
# ----------------------------------------------------------------------------


if __name__ == "__main__":
    import PyQt5_ApplicationClass
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('Fusion')  # 'Fusion' ... QtWidgets.QStyle
    window = PyQt5_ApplicationClass.AppWindow()
    window.setWindowTitle("GyroVibroTest")
    # window.resize(850, 500)
    window.show()
    sys.exit(app.exec())


#    def plot_fft_median(self, freq_data: np.ndarray, special_points: np.ndarray, folder: str):
#         self.logger.info("Final median plot")
#         # self.append_fft_plot_tab()
#         self.last_tab_layout = QtWidgets.QGridLayout(spacing=0)
#         self.last_tab_layout.addWidget(self.groupbox, 0, 1, 2, 1) 

#         index = self.count() - 1
#         self.tab_widget_page_list.append(QtWidgets.QWidget(self))
#         self.append_amp_plot()
#         self.last_tab_layout.addWidget(self.amp_plot_list[index], 0, 0, 1, 1)
#         self.append_phase_plot()
#         self.last_tab_layout.addWidget(self.phase_plot_list[index], 1, 0, 1, 1)
#         self.tab_widget_page_list[-1].setLayout(self.last_tab_layout)
#         self.addTab(
#             self.tab_widget_page_list[-1], f"ЧХ &{index + 1}")  # FC 
#         self.setTabText(self.count() - 1, "&АФЧХ (средний)")  # FC average
#         self.setCurrentIndex(self.count() - 1)
  
#         # self.del_btn.clicked.connect(
#         #     self.text_())
#         self.name_info.setText(folder)

#         self.x = freq_data[:, -4]  # [1, 5, 13, 20, 55, 62, 80]
#         # # amp =np.array([96, 97, 108, 112, 113, 114, 121, 127, 136, 163, 166, 191, 219, 240, 258, 284, 400, 371, 450, 699, 791, 1154, 631, 697, 722, 743, 834, 823, 918, 995, 1022, 1125, 1220, 1244, 1373, 1468, 1618, 1851, 1865, 2166, 2548, 2539, 3409, 3521, 3514, 4220, 3473, 2573, 3081, 3028, 3028, 3230, 3056, 3132, 3102, 3621, 3669, 4561])/100
#         self.y = freq_data[:, -3]  # np.array([1, 1.2, 1.3, 1.5, 5, .5, .35])
#         # self.amp_curves[-1].setData(self.x, self.y)
#         # self.plot_2d_scatter(self.time_plot, self.x, self.y)
#         # self.cursor = Qt.BlankCursor
#         # self.amp_plot_list[-1].setCursor(self.cursor)  # self.cursor ##########################

#         # Add lines
#         # self.amp_plot_list[-1].addItem(self.infinite_line_x, ignoreBounds=True)
#         # self.amp_plot_list[-1].addItem(self.infinite_line_y, ignoreBounds=True)
    
#         # self.amp_plot_list[-1].addItem(self.cursorlabel)#################################
#         # self.proxy = pg.SignalProxy(
#         #     self.amp_plot_list[-1].scene().sigMouseMoved, delay=0.15,##############################
#         #     rateLimit=15, slot=self.update_crosshair)

#         # self.proxy = pg.SignalProxy(
#             # self.phase_plot_list[-1].scene().sigMouseMoved, delay=0.15,
#             # rateLimit=15, slot=self.update_crosshair) ####################
#         self.mouse_x = None
#         self.mouse_y = None
#         # # self.plot_fft(True)
#         for i in range(self.GYRO_NUMBER):
#             self.amp_curves[-1 - i].setData(freq_data[:, -4],
#                                         freq_data[:, -3])
#             self.phase_curves[-1 - i].setData(freq_data[:, -4],
#                                         freq_data[:, -2])
#         self.amp_plot_list[-1].autoRange()
#     # def plot_2d_scatter(self, plot, x, y, color=(66, 245, 72)):
#     #     brush = pg.mkBrush(color)
#     #     scatter = pg.ScatterPlotItem(size=5, brush=brush)
#     #     scatter.addPoints(x, y)
#     #     plot.addItem(scatter)