# import numpy as np
# import pandas as pd
# import os
from openpyxl import load_workbook

# file_path = 'fff.xlsx'
# wb = load_workbook(file_path)
# ws = wb['ПУС X']  # or wb.active
# # ws = wb.active  # or wb.active
# ws['G6'] = 'g'
# # print(ws[f'G6'].value is None)
# print(ws[f'F6'].value is None)
# ws['G7'] = 1523
# ws['G6'] = f'=E5+1500/$D$1-1500/$D$2'
wb = load_workbook("D:/Gyro2023_Git/БИЧЭ-К_№157.xlsx")
ws = wb['ПУС Z']  # or 
ws['F4'] = '--- задержка в блоке, мс 0000'
ws.column_dimensions['F'].width = 51
try:
    wb.save("D:/Gyro2023_Git/БИЧЭ-К_№157.xlsx")
except IOError:
    print("Error!")
# --------------------------------------------------------------
# sheets = ['ПУС X', 'ПУС Y', 'ПУС Z']
# for j in range(1, 168 + 1):
#     file_path = f'D:/БИЧЭ-К_АФЧХ_2023_/БИЧЭ-К_№{j}.xlsx'
    
#     wb = load_workbook(file_path)

#     for sheet_name in sheets:
#         ws = wb[sheet_name]  # or wb.active
#         ws['F4'] = 'задержка в блоке, мс'
#         ws.column_dimensions['F'].width = 21
#         # for i in range(5, 50):
#         #     if ws[f'F{i}'].value is None:
#         #         continue
#         #     ws[f'F{i}'] = f'=E{i}-1500/$D$1+1500/$D$2'
#     wb.save(file_path)

# --------------------------------------------------------------
# sheets = ['ПУС X', 'ПУС Y', 'ПУС Z']
# for j in range(104, 168 + 1):
#     file_path = f'D:/БИЧЭ-К_АФЧХ_2023/БИЧЭ-К_№{j}.xlsx'
#     wb = load_workbook(file_path)

#     for sheet_name in sheets:
#         ws = wb[sheet_name]  # or wb.active
#         ws[f'D1'] = 1000
#         for i in range(5, 50):
#             if ws[f'F{i}'].value is None:
#                 continue
#             ws[f'F{i}'] = f'=E{i}-1500/$D$1+1500/$D$2'
#     wb.save(file_path)
# --------------------------------------------------------------
# sheets = ['ПУС X', 'ПУС Y', 'ПУС Z']
# list_ = [83, 90, 91, 94, 98, 101, 102]
# for j in list_:
#     file_path = f'D:/БИЧЭ-К_АФЧХ_2023/БИЧЭ-К_№{j}.xlsx'
#     wb = load_workbook(file_path)

#     for sheet_name in sheets:
#         ws = wb[sheet_name]  # or wb.active
#         ws[f'D1'] = 1000
#         for i in range(5, 50):
#             if ws[f'F{i}'].value is None:
#                 continue
#             ws[f'F{i}'] = f'=E{i}-1500/$D$1+1500/$D$2'
#     wb.save(file_path)
# --------------------------------------------------------------
# sheets = ['ПУС Z',
#           'ПУС Y',
#           'ПУС Z',
#           'ПУС Z',
#           'ПУС Z'
#           ]
# list_ = [7, 9, 16, 18, 21]
# k = 0
# for j in list_:
#     file_path = f'D:/БИЧЭ-К_АФЧХ_2023/БИЧЭ-К_№{j}.xlsx'
#     wb = load_workbook(file_path)

#     sheet_name = sheets[k]
#     k+=1
#     ws = wb[sheet_name]  # or wb.active
#     ws[f'D1'] = 741
#     for i in range(5, 50):
#         if ws[f'F{i}'].value is None:
#             continue
#         ws[f'F{i}'] = f'=E{i}-1500/$D$1+1500/$D$2'
#     wb.save(file_path)
# --------------------------------------------------------------
# sheets = ['ПУС X',
#           'ПУС Y',
#           'ПУС X',
#           'ПУС Y',
#           'ПУС Y',
#           'ПУС Z',
#           'ПУС Z',
#           'ПУС Y'
#           ]
# list_ = [13,
#          17,
#          44,
#          69,
#          76,
#          78,
#          85,
#          95
#          ]
# k = 0
# for j in list_:
#     file_path = f'D:/БИЧЭ-К_АФЧХ_2023/БИЧЭ-К_№{j}.xlsx'
#     # file_path = "БИЧЭ-К_№157.xlsx"
#     wb = load_workbook(file_path)

#     sheet_name = sheets[k]
#     k+=1
#     ws = wb[sheet_name]  # or wb.active
#     ws[f'D1'] = 1000
#     for i in range(5, 50):
#         if ws[f'F{i}'].value is None:
#             continue
#         ws[f'F{i}'] = f'=E{i}-1500/$D$1+1500/$D$2'
#     wb.save(file_path)
# --------------------------------------------------------------
# sheets = [['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z'],
#           ['ПУС Y', 'ПУС Z']
#           ]
# print(len(sheets))
# list_ = [80,
#          81,
#          82,
#          84,
#          86,
#          87,
#          88,
#          89,
#          92,
#          93,
#          96,
#          97,
#          99,
#          100,
#          103
#          ]
# print(len(list_))
# # print(sheets[0])
# k = 0
# for j in list_:
#     file_path = f'D:/БИЧЭ-К_АФЧХ_2023/БИЧЭ-К_№{j}.xlsx'
#     wb = load_workbook(file_path)

#     for sheet_name in sheets[k]:
#         ws = wb[sheet_name]  # or wb.active
#         ws[f'D1'] = 1000
#         for i in range(5, 50):
#             if ws[f'F{i}'].value is None:
#                 continue
#             ws[f'F{i}'] = f'=E{i}-1500/$D$1+1500/$D$2'
#     k += 1
#     wb.save(file_path)
