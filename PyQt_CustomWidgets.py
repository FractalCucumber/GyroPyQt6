import logging
import re
import sys
import numpy as np
from PyQt5 import QtWidgets, QtCore, QtGui
import pyqtgraph.exporters
import pyqtgraph as pg
from openpyxl import load_workbook  # лучше использовать Pandas, наверняка работает быстрее
# from typing import overload
# from PyQt_Functions import get_icon_by_name, get_res_path
import PyQt_ProjectsComboBox
from PyQt_Functions import check_name_simple
import os


class CustomViewBox(pg.ViewBox):
    def __init__(self, *args, **kwds):
        kwds['enableMenu'] = True
        # kwds['enableMenu'] = False
        pg.ViewBox.__init__(self, *args, **kwds)
        # self.setMouseMode(self.RectMode)

    #  reimplement right-click to zoom out
    def mouseClickEvent(self, ev):
        if ev.button() == QtCore.Qt.MouseButton.RightButton:
            self.autoRange()

    # #  reimplement mouseDragEvent to disable continuous axis zoom
    # def mouseDragEvent(self, ev, axis=None):
    #     if axis is not None and ev.button() == QtCore.Qt.MouseButton.RightButton:
    #         ev.ignore()
    #     else:
    #         pg.ViewBox.mouseDragEvent(self, ev, axis=axis)


class CustomTabWidget(QtWidgets.QTabWidget):
    warning_signal = QtCore.pyqtSignal(str)
    get_filename_signal = QtCore.pyqtSignal(bool)

    def __init__(self, GYRO_NUMBER, fs=1000, logger_name='', parent=None):
        # QtWidgets.QTabWidget.__init__(self)
        super(CustomTabWidget, self).__init__(parent)
        self.GYRO_NUMBER = GYRO_NUMBER
        self.visibility_flags = [True] * (self.GYRO_NUMBER + 1)
        self.LABEL_STYLE = {'color': '#FFF', 'font-size': '16px'}
        self.COLOR_LIST = ['r', 'g', '#006bf7']
        self.fs = fs
        self.selected_files_to_fft = []
        self.pt = 12
        # self.settings = QtCore.QSettings("settings")
        self.logger = logging.getLogger(logger_name)
        pg.setConfigOption('background', '#151515')
        pg.setConfigOption('foreground', '#BBBBBB')
# ------ time plot ------------------------------------------------------------

        self.time_plot_item = pg.PlotItem(viewBox=CustomViewBox())
        self.time_plot = pg.PlotWidget(plotItem=self.time_plot_item)
        self.time_plot.setLimits(xMin=-0.01)
        self.time_plot_item.setTitle('Угловая скорость', size=f'{self.pt}pt')  # Velosity Graph
        self.time_plot_item.showGrid(x=True, y=True)
        self.time_plot_item.addLegend(offset=(-1, 1), labelTextSize=f'{self.pt}pt',
                                      labelTextColor=pg.mkColor('w'))
        self.time_plot_item.setLabel('left', 'Velosity',
                                     units='\u00b0/second', **self.LABEL_STYLE)
        self.time_plot_item.setLabel('bottom', 'Time',
                                     units='seconds', **self.LABEL_STYLE)

        self.time_curves = [self.time_plot_item.plot(pen='w', name="encoder")]
        for i in range(self.GYRO_NUMBER):
            self.time_curves.append(self.time_plot_item.plot(
                pen=self.COLOR_LIST[i], name=f"gyro {i + 1}"))

        self.region = pg.LinearRegionItem([0, 0], movable=False)
        self.time_plot_item.addItem(self.region)

        # self.x = [1, 5, 10, 15, 20, 25, 33, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 100, 102, 105, 107, 110, 115, 118, 120, 122, 125, 130, 135, 140, 150, 156, 162, 170, 180, 190, 200, 205, 210, 215, 220, 225, 230, 235, 240, 243, 247, 250, 255, 260, 265, 270, 275, 280, 285, 290, 295, 300, 305, 310, 315]
        # self.y =np.array(
        #     [94, 99, 105, 114, 112, 113, 122, 127, 137, 160, 169, 205, 221, 243, 292, 306, 339, 392, 419, 554, 861, 913, 1079, 1276, 595, 645, 659, 666, 781, 810, 863, 948, 1024, 1103, 1227, 1329, 1319, 1362, 1468, 1791, 1921, 1959, 2641, 2477, 3384, 3484, 3254, 4482, 3888, 3931, 2795, 3021, 3246, 3046, 3147, 3505, 3424, 3430, 4049, 4157, 4385, 3980, 5340, 4897]
        #     )/100
        # self.time_curves[0].setData(self.x, self.y)
# ------ Tab widget -----------------------------------------------------------
        self.page = QtWidgets.QWidget(self)
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.time_plot)
        self.spectrum_button = QtWidgets.QPushButton("От времени")  # Time plot
        layout.addWidget(self.spectrum_button)
        self.spectrum_button.clicked.connect(self.switch_plot_x_axis)
        self.page.setLayout(layout)
        self.addTab(self.page, "\u03C9(t)")  # &Time plot От времени
        self.tab_widget_page_list: list[QtWidgets.QWidget] = []
        self.phase_curves: list[pg.PlotCurveItem] = []
        self.amp_curves: list[pg.PlotCurveItem] = []
        self.phase_plot_list: list[pg.PlotWidget] = []
        self.amp_plot_list: list[pg.PlotWidget] = []

# -----------------------------------------------------------------------------
        self.groupbox = QtWidgets.QGroupBox(
            '', maximumWidth=190, minimumWidth=150)
        self.median_plot_groupbox_layout = QtWidgets.QGridLayout(spacing=6)
        # self.median_plot_groupbox_layout.setRowStretch(4, 0)
        # self.median_plot_groupbox_layout.setSpacing(0)
        # self.median_plot_groupbox_layout.setContentsMargins(0, 0, 0, 0)
        self.groupbox.setLayout(self.median_plot_groupbox_layout)

        self.groupbox_list: list[QtWidgets.QGroupBox] = []
        self.amp_info_line_edit_list: list[QtWidgets.QLineEdit] = []
        self.freq_info_line_edit_list: list[QtWidgets.QLineEdit] = []
        self.sensor_name_line_edit_list: list[QtWidgets.QLineEdit] = []

        for i in range(self.GYRO_NUMBER):
            self.append_gyro_groupbox()
            self.median_plot_groupbox_layout.addWidget(
                self.groupbox_list[-1], 2 + i, 0, 1, 3)

        self.projects_combo_box = PyQt_ProjectsComboBox.ProjectsComboBox()
        self.median_plot_groupbox_layout.addWidget(
            self.projects_combo_box, 0, 0, 1, 3)

        self.ok_btn = QtWidgets.QPushButton("Запись в Excel")
        self.median_plot_groupbox_layout.addWidget(self.ok_btn, 1, 0, 1, 3)
        self.ok_btn.clicked.connect(self.write_xlsx)

        self.get_filenames_btn = QtWidgets.QPushButton("АФЧХ для файла")
        self.median_plot_groupbox_layout.addWidget(
            self.get_filenames_btn, 7, 0, 1, 3)
        self.get_filenames_btn.clicked.connect(self.get_filenames)

        self.plot_fft_median()

        self.selected_curve = 0
        self.groupbox_list[self.selected_curve].setChecked(True)
        for j in range(1, len(self.groupbox_list)):
            self.groupbox_list[j].setChecked(False)
        for i in range(2):
            self.infinite_x_line_list[i].setPen(
                pg.mkPen(self.COLOR_LIST[self.selected_curve]))
        # self.groupbox_clicked(self.groupbox_list[0])
        # print(self.palette().window().color().name())
        # pg.setConfigOption('background', self.palette().window().color().name())
# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------
#
################################################################################################
#
# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------
# ----- plotting -------------------------------------------------------------------------------
            
    # @overload
    def plot_time_graph(self, time: np.ndarray,
                        enc_data: np.ndarray, gyro_data: list):
        self.time_curves[0].setData(time, enc_data)
        for i in range(self.GYRO_NUMBER):
            self.time_curves[i + 1].setData(time, gyro_data[i])

    # @overload
    # def plot_time_graph(self, time_data: np.ndarray,
    #                     enc_data: np.ndarray, gyro_data: np.ndarray):
    #     self.time_curves[0].setData(time_data, enc_data)
    #     for i in range(self.GYRO_NUMBER):
    #         self.time_curves[i + 1].setData(time_data, gyro_data)

    def set_fft_data(self, freq_data: np.ndarray, frame: list):
        """Adds points to frequency graphs"""
        # self.logger.info(f"plot_fft, len {len(self.amp_curves)}")
        # self.logger.info(f"plot_fft, len amp_plot_list {len(self.amp_plot_list)}")
        # self.logger.info(f"self.amp_plot_list {self.amp_plot_list}")
        # self.logger.info(f"self.amp_curves {self.amp_curves}")
        # self.logger.info(f"self.amp_curves {self.amp_plot_list}")
        # self.logger.info(
            # f"self.amp_plot_list[-1].getPlotItem().curves {self.amp_plot_list[-1].getPlotItem().curves[-1]}")
        # до этого момента менются данные
        # for i in range(len(self.amp_curves)):
        #     print(f"{self.amp_curves[i]}")
        #     print(f"{i} {self.amp_curves[i].getData()}")
        for i in range(self.GYRO_NUMBER):
            # self.logger.info(f"getPlotItem().curves[i] {freq_data[:, 0, i]}")
            # self.amp_plot_list[-1].getPlotItem().curves[i].setData(freq_data[:, 0, i],
                                        # freq_data[:, 1, i])
            # self.phase_plot_list[-1].getPlotItem().curves[i].setData(freq_data[:, 0, i],
                                        # freq_data[:, 2, i])
            # ind = -1 - i
            ind = self.GYRO_NUMBER * (self.count() - 2) + i
            # self.logger.info(f"ind {ind}")
            # self.logger.info(f"self.amp_curves {self.amp_curves[ind]}")
            # self.logger.info(f"self.phase_curves {self.phase_curves[ind]}")
            
            self.amp_curves[ind].setData(np.copy(freq_data[:, 0, i]),
                                         np.copy(freq_data[:, 1, i]))
            self.phase_curves[ind].setData(np.copy(freq_data[:, 0, i]),
                                           np.copy(freq_data[:, 2, i]))
        # for i in range(len(self.amp_curves)):
        #     print(f"after {self.amp_curves[i]}")
        #     print(f"{i} after {self.amp_curves[i].getData()}")
        self.region.setRegion([frame[0]/self.fs, frame[1]/self.fs])
        self.amp_plot_list[-1].autoRange()
        # self.phase_plot_list[-1].autoRange()

# ------ FFT median plot ------------------------------------------------------------------
    def set_fft_median_data(self, freq_data: np.ndarray,
                            special_points: np.ndarray, folder: list):
        self.logger.info(f"Final median plot, sensor: {folder}")
        temp = np.empty(
            (freq_data.shape[0] + 1, freq_data.shape[1], freq_data.shape[2]))
        # for i in range(self.GYRO_NUMBER):
        for i in range(len(self.groupbox_list)):
            self.groupbox_list[i].setVisible(False)
        # for i in range(freq_data.shape[2]):  # !!!
        for i in range(len(folder)):
            self.groupbox_list[i].setVisible(True)
            temp[:, :, i] = np.insert(
                freq_data[:, :, i],
                int(special_points[-1, i]), special_points[:-1, i], axis=0)

            self.amp_curves[i].setData(temp[:, -4, i],
                                        temp[:, -3, i])
            self.phase_curves[i].setData(temp[:, -4, i],
                                        temp[:, -2, i])
            self.amp_info_line_edit_list[i].setText(
                f"{np.max(freq_data[:, -3, i]):.3f}")
                # str(round(np.max(freq_data[:, -3, i]), 3))) #############################################################
            self.freq_info_line_edit_list[i].setText(
                f"{freq_data[np.argmax(freq_data[:, -3, i]), -4, i]:.2f}")
                # str(round(freq_data[np.argmax(freq_data[:, -3, i]), -4, i], 2)))
            self.sensor_name_line_edit_list[i].setText(folder[i])
        # self.freq_data = freq_data
        self.freq_data = temp
        self.amp_plot_list[0].autoRange()
        self.phase_plot_list[0].autoRange()
        self.mouse_x = None
        self.mouse_y = None

# ----- cursor ---------------------------------------------------------------------------------
    def update_crosshair(self, e):
        pos = e[0]
        if (self.amp_plot_list[0].plotItem.sceneBoundingRect().contains(pos)
            and self.freq_data[0, 0, 0] != -1):
            mouse_point = self.amp_plot_list[0].plotItem.vb.mapSceneToView(pos)
            # self.selected_curve
            index = np.nanargmin(
                np.abs(self.freq_data[:, -4, self.selected_curve] - mouse_point.x()))
            if 0 <= index < len(self.freq_data[:, -4, self.selected_curve]):  # index >= 0 and 
                self.cursor_label.setText(
                    f"f {self.freq_data[index, -4, self.selected_curve]:.2f}\n" +
                    f"A {self.freq_data[index, -3, self.selected_curve]:.2f}\n" +
                    f"\u03C6 {self.freq_data[index, -2, self.selected_curve]:.1f}")
                # map(pg.InfiniteLine.setPos, self.infinite_x_line_list, [self.freq_data[index, -4, self.selected_curve] * 2])
                for i in range(2):
                    self.infinite_x_line_list[i].setPos(
                        self.freq_data[index, -4, self.selected_curve]) 
                self.mouse_x = (self.freq_data[index, -4, self.selected_curve]) # freq_data[:, -4]
                self.mouse_y = (self.freq_data[index, -3, self.selected_curve]) # freq_data[:, -3]
            
    def mouseDoubleClickEvent(self, e):
        if (e.buttons() & QtCore.Qt.LeftButton & (
            self.amp_plot_list[0].underMouse() |
            self.phase_plot_list[0].underMouse()
            )) and self.freq_data[0, 0, 0] != -1:
            # )) and self.freq_data.size:
            self.amp_info_line_edit_list[self.selected_curve].setText(
                f'{self.mouse_y:.3f}')
            self.freq_info_line_edit_list[self.selected_curve].setText(
                f'{self.mouse_x:.2f}')
            self.cursor_label.setPos(self.mouse_x + 3, self.mouse_y + 0.3)
# ----- Excel and file selecting ----------------------------------------------

    @QtCore.pyqtSlot()
    def get_filenames(self):
        self.selected_files_to_fft, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self,
            "Выберите файлы для построения АФЧХ",
            ".", "Text Files (*.txt)")
            # "All Files (*);;Python Files (*.py);;Text Files (*.txt)")
        if len(self.selected_files_to_fft):
            self.get_filename_signal.emit(True)
 
    @QtCore.pyqtSlot()
    def write_xlsx(self):
        return
        self.currnt_xlsx_path = \
            self.projects_combo_box.projects_dict[
                self.projects_combo_box.currentText()]
        self.currnt_xlsx_path = "test.xlsx"  ########################################
        try:
            wb = load_workbook(self.currnt_xlsx_path)
        except FileNotFoundError:
            self.warning_signal.emit("File not found!")
            return
        ws = wb['Лист1']  # or wb.active
        ws['F3'] = self.sensor_name_line_edit_list[self.selected_curve].text()
        ws['F4'] = self.freq_info_line_edit_list[self.selected_curve].text()
        ws['F5'] = self.amp_info_line_edit_list[self.selected_curve].text()
        # ws.column_dimensions['F'].width = 51
        try:
            wb.save(self.currnt_xlsx_path)
        except IOError:
            self.warning_signal.emit("File was open! Close and try again")

# ----- plot change -----------------------------------------------------------

    @QtCore.pyqtSlot()
    def groupbox_clicked(self):
        # for i in range(len(self.groupbox_list)):
        for i in range(self.freq_data.shape[2]):
            if self.sender() == self.groupbox_list[i]:
                flag = False
                break
        else:
            flag = True

        for j in range(len(self.groupbox_list)):
            self.groupbox_list[j].setChecked(False)
        if self.freq_data[0, 0, 0] == -1 or flag:
            return

        self.selected_curve = i
        self.groupbox_list[i].setChecked(True)
        if self.groupbox_list[i].isChecked():
            for j in range(2):
                self.infinite_x_line_list[j].setPen(pg.mkPen(self.COLOR_LIST[i]))

    def clear_plots(self):
        self.freq_data = np.ndarray((1, 4, 1))  # лучше пустой массив создавать
        self.freq_data.fill(-1)
        self.region.setRegion([0, 0])
        self.logger.info(f"len amp prev {len(self.amp_curves)}")
        self.logger.info(f"len phase_curves prev {len(self.phase_curves)}")
        for i in range(self.count() - 2):
            for _ in range(self.GYRO_NUMBER):
                self.amp_curves.pop()
                self.phase_curves.pop()
            self.amp_plot_list.pop()
            self.phase_plot_list.pop()
            self.removeTab(2)
            self.tab_widget_page_list.pop()
        #         phase_plot_list
        self.logger.info(f"len amp {len(self.amp_curves)}")
        self.logger.info(f"len phase_curves {len(self.phase_curves)}")
        self.time_curves[0].setData([])
        for i in range(self.GYRO_NUMBER):
            self.time_curves[i + 1].setData([])
            self.amp_curves[i].setData([])
            self.phase_curves[i].setData([])
        # self.phase_curves: list[pg.PlotCurveItem] = []
        # self.amp_curves: list[pg.PlotCurveItem] = []
        # self.phase_plot_list: list[pg.PlotWidget] = []
        # self.amp_plot_list: list[pg.PlotWidget] = []
        # self.tab_widget_page_list: list[QtWidgets.QWidget] = []

    @QtCore.pyqtSlot()
    def change_curve_visibility(self):
        ind = int(self.sender().objectName())
        # self.flag_visibility[num] = self.check_box_list[num].checkState()
        self.visibility_flags[ind] = not self.visibility_flags[ind]
        self.time_curves[ind].setVisible(self.visibility_flags[ind])
        if ind == 0:
            return
        for i in range(self.count() - 1):
            self.phase_curves[i*self.GYRO_NUMBER - 1 + ind].setVisible(
                self.visibility_flags[ind])
            self.amp_curves[i*self.GYRO_NUMBER - 1 + ind].setVisible(
                self.visibility_flags[ind])

    @QtCore.pyqtSlot()
    def switch_plot_x_axis(self):
        """
        Switching between time plot and spectrum
        """
        if self.spectrum_button.text() == "Спектр":  # Frequency plot
            self.spectrum_button.setText("От времени")  # Time plot
            self.time_plot_item.ctrl.fftCheck.setChecked(False)
            self.time_plot_item.setLabel(
                'bottom', 'Time', units='seconds')
            self.region.show()
        else:
            self.spectrum_button.setText("Спектр")  # Frequency plot
            self.time_plot_item.ctrl.fftCheck.setChecked(True)
            self.time_plot_item.setLabel(
                'bottom', 'Frequency', units='Hz')
            self.region.hide()

    @QtCore.pyqtSlot(str)
    def save_plot_image(self, path: str):
        if not os.path.isdir(os.path.dirname(path)):
            self.warning_signal.emit(f"Folder {os.path.dirname(path)} doesn't exist!")
            return
        pyqtgraph.exporters.ImageExporter(
            self.time_plot_item).export(
                check_name_simple(path + '_time_plot.png'))
        for i in range(self.count() - 1):
            pyqtgraph.exporters.ImageExporter(
                self.amp_plot_list[i].getPlotItem()).export(
                    check_name_simple(path + f'_amp_plot_{i + 1}.png'))
            pyqtgraph.exporters.ImageExporter(
                self.phase_plot_list[i].getPlotItem()).export(
                    check_name_simple(path + f'_phase_plot_{i + 1}.png'))

    def plot_fft_median(self):
        self.freq_data = np.ndarray((1, 4, 1))  # лучше пустой массив создавать
        self.freq_data.fill(-1)
        self.last_tab_layout = QtWidgets.QGridLayout(spacing=0)
        self.last_tab_layout.addWidget(self.groupbox, 0, 1, 2, 1) 

        self.tab_widget_page_list.append(QtWidgets.QWidget(self))
        self.append_amp_plot()
        self.last_tab_layout.addWidget(self.amp_plot_list[0], 0, 0, 1, 1)
        self.append_phase_plot()
        self.last_tab_layout.addWidget(self.phase_plot_list[0], 1, 0, 1, 1)
        self.tab_widget_page_list[0].setLayout(self.last_tab_layout)
        self.addTab(
            self.tab_widget_page_list[0], "&АФЧХ (средний)")  # FC average
  
        # # amp =np.array([96, 97, 108, 112, 113, 114, 121, 127, 136, 163, 166, 191, 219, 240, 258, 284, 400, 371, 450, 699, 791, 1154, 631, 697, 722, 743, 834, 823, 918, 995, 1022, 1125, 1220, 1244, 1373, 1468, 1618, 1851, 1865, 2166, 2548, 2539, 3409, 3521, 3514, 4220, 3473, 2573, 3081, 3028, 3028, 3230, 3056, 3132, 3102, 3621, 3669, 4561])/100
        # self.cursor = QtCore.Qt.BlankCursor
        self.amp_plot_cursor = QtCore.Qt.CrossCursor # was
        self.amp_plot_list[0].setCursor(self.amp_plot_cursor)  # self.cursor ##########################
        self.phase_plot_cursor = QtCore.Qt.CrossCursor # was
        self.phase_plot_list[0].setCursor(self.phase_plot_cursor)  # self.cursor ##########################

        self.infinite_x_line_list = [pg.InfiniteLine(angle=90, pen=pg.mkPen(self.COLOR_LIST[0]))
                                     for _ in range(2)]
        self.amp_plot_list[0].addItem(self.infinite_x_line_list[0], ignoreBounds=True)
        self.phase_plot_list[0].addItem(self.infinite_x_line_list[1], ignoreBounds=True)
        self.phase_plot_list[0].addItem(
            pg.InfiniteLine(
                pos=-180, angle=0, movable=False,
                pen=pg.mkPen((205, 205, 205), dash=[10, 2], width=0.75)),
                ignoreBounds=True)  # -180 degrees
        self.cursor_label = pg.TextItem()
        self.cursor_label.setPos(3, 0.8)

        self.amp_plot_list[0].addItem(self.cursor_label) #################################
        self.proxy_amp = pg.SignalProxy(
            self.amp_plot_list[0].scene().sigMouseMoved, delay=0.1, ##############################
            rateLimit=12, slot=self.update_crosshair)
        self.proxy_phase = pg.SignalProxy(
            self.phase_plot_list[0].scene().sigMouseMoved, delay=0.1, ##############################
            rateLimit=12, slot=self.update_crosshair)
        self.amp_plot_list[0].autoRange()

    def append_fft_plot_tab(self):
        """
        Create new tab and append amplitude and grequency graphs
        """
        index = self.count() - 1
        self.tab_widget_page_list.append(QtWidgets.QWidget(self))
        layout = QtWidgets.QVBoxLayout(spacing=0)
        self.append_amp_plot()
        # layout.addWidget(self.amp_plot_list[index])
        layout.addWidget(self.amp_plot_list[-1])
        self.append_phase_plot()
        # layout.addWidget(self.phase_plot_list[index])
        layout.addWidget(self.phase_plot_list[-1])
        self.tab_widget_page_list[-1].setLayout(layout)
        self.addTab(
            self.tab_widget_page_list[-1], f"ЧХ &{index}")  # FC        

    def append_amp_plot(self):
        amp_plot_item = pg.PlotItem(viewBox=CustomViewBox(),
                                         name=f'amp_plot{self.count()}')
        amp_plot_item.setXLink(f'phase_plot{self.count()}')
        amp_plot_item.setTitle('АФЧХ', size=f'{self.pt - 1}pt')  # Amp Graph
        amp_plot_item.showGrid(x=True, y=True)
        amp_plot_item.addLegend(offset=(-1, 1), labelTextSize=f'{self.pt - 1}pt',
                                     labelTextColor=pg.mkColor('w'))
        amp_plot_item.setLabel('left', 'Amplitude',
                                    units="", **self.LABEL_STYLE)
        # amp_plot_item.setLabel('bottom', 'Frequency',
        #                             units='Hz', **self.LABEL_STYLE)
        # self.SYMBOL_SIZE = 6
        for i in range(self.GYRO_NUMBER):
            self.amp_curves.append(amp_plot_item.plot(
                pen=self.COLOR_LIST[i], name=f"gyro {i + 1}", symbol="o",
                symbolSize=6, symbolBrush=self.COLOR_LIST[i]))
        self.amp_plot_list.append(pg.PlotWidget(plotItem=amp_plot_item))
        self.amp_plot_list[-1].getAxis('left').setWidth(60)
        self.amp_plot_list[-1].setLimits(
            xMin=-4, xMax=int(self.fs * 0.53), yMin=-0.08, yMax=100)

    def append_phase_plot(self):
        phase_plot_item = pg.PlotItem(viewBox=CustomViewBox(),
                                           name=f'phase_plot{self.count()}')
        phase_plot_item.setXLink(f'amp_plot{self.count()}')
        phase_plot_item.setTitle('ФЧХ', size='12pt')  # Phase Graph
        phase_plot_item.showGrid(x=True, y=True)
        phase_plot_item.addLegend(offset=(-1, 1), labelTextSize=f'{self.pt - 1}pt',
                                       labelTextColor=pg.mkColor('w'))
        phase_plot_item.setLabel('left', 'Phase',
                                      units='degrees', **self.LABEL_STYLE)  # rad
        phase_plot_item.setLabel('bottom', 'Frequency',
                                      units='Hz', **self.LABEL_STYLE) # \u00b0
        for i in range(self.GYRO_NUMBER):
            self.phase_curves.append(phase_plot_item.plot(
                pen=self.COLOR_LIST[i], name=f"gyro {i + 1}", symbol="o",
                symbolSize=6, symbolBrush=self.COLOR_LIST[i]))
        self.phase_plot_list.append(
            pg.PlotWidget(plotItem=phase_plot_item))
        self.phase_plot_list[-1].getAxis('left').setWidth(60)
        self.phase_plot_list[-1].setLimits(
            xMin=-4, xMax=int(self.fs * 0.53), yMin=-375, yMax=20)

    def append_gyro_groupbox(self):
        ind = len(self.groupbox_list)
        self.groupbox_list.append(QtWidgets.QGroupBox(
            f'gyro {ind + 1}', maximumWidth=190, minimumWidth=150, maximumHeight=290,
            checkable=True, objectName='small'))
        self.groupbox_layout = QtWidgets.QGridLayout(spacing=2)
        self.groupbox_list[-1].clicked.connect(self.groupbox_clicked)
        # self.median_plot_groupbox_layout.setRowStretch(4, 0)
        # self.median_plot_groupbox_layout.setContentsMargins(0, 0, 0, 0)
        self.groupbox_list[-1].setLayout(self.groupbox_layout)
        # QToolButton # есть wordWrap
        amp_info_label = QtWidgets.QLabel("amp")
        self.groupbox_layout.addWidget(
            amp_info_label, 0, 0, 1, 1)
        self.amp_info_line_edit_list.append(QtWidgets.QLineEdit())
        self.groupbox_layout.addWidget(
            self.amp_info_line_edit_list[-1], 0, 1, 1, 2)

        freq_info_label = QtWidgets.QLabel("f, Hz")
        self.groupbox_layout.addWidget(
            freq_info_label, 1, 0, 1, 1)
        self.freq_info_line_edit_list.append(QtWidgets.QLineEdit())
        self.groupbox_layout.addWidget(
            self.freq_info_line_edit_list[-1], 1, 1, 1, 2)

        sensor_name_info_label = QtWidgets.QLabel(
            "номер:", maximumHeight=40)
        self.groupbox_layout.addWidget(
            sensor_name_info_label, 2, 0, 1, 1)
        self.sensor_name_line_edit_list.append(QtWidgets.QLineEdit())
        self.groupbox_layout.addWidget(
            self.sensor_name_line_edit_list[-1], 2, 1, 1, 2)

# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
#
###############################################################################
#
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------


class CustomComboBox(QtWidgets.QComboBox):
    def __init__(self,
                 settings: QtCore.QSettings,  # передавать сразу объект QtCore.QSettings, чтобы не создавать их
                 settings_name: str = 'Deafault',
                 default_items_list = [''],
                 editable_flag=True,
                 uint_validator_enable=True,
                 parent=None):
        super(CustomComboBox, self).__init__(parent)
        self.settings_name = settings_name
        self.setEditable(True)
        self.lineEdit().setReadOnly(not editable_flag)
        self.lineEdit().setAlignment(
            QtCore.Qt.AlignmentFlag.AlignCenter)
        if uint_validator_enable:
            self.int_validator = QtGui.QIntValidator(bottom=0)
            self.setValidator(self.int_validator)

        self.currentTextChanged.connect(
            lambda value: self.setItemText(self.currentIndex(), value))

        # self.settings = QtCore.QSettings(settings_name)
        self.settings = settings
        if self.settings.contains("item_" + self.settings_name):
            for i in range(self.count()):
                self.removeItem(i)
            self.addItems(
                self.settings.value("item_" + self.settings_name))
        else:
            if not len(default_items_list[0]):
                return
            self.addItems(default_items_list)
        if self.settings.contains("curr_index_" + self.settings_name):
            if self.count() >= self.settings.value("curr_index_" + self.settings_name):
                self.setCurrentIndex(
                    self.settings.value("curr_index_" + self.settings_name))

        # self.save_value = lambda: self.settings.setValue(
        #     "items",
        #     [self.itemText(i) for i in range(self.count())])
        # self.save_index = lambda: self.settings.setValue(
        #     "curr_index", self.currentIndex())
        # self.save_name = lambda: self.settings.setValue(
        #     "COM_current_name", self.currentText())
    def get_ind(self):
        if self.settings.contains("name" + self.settings_name):
            for i in range(self.count()):
                if self.itemText(i) == self.settings.value("name" + self.settings_name):
                    self.setCurrentIndex(i)
                    break

    def save_all(self):
        self.save_value()
        self.save_index()

    def save_value(self):
        if self.count():
            self.settings.setValue(
                "item_" + self.settings_name,
                [self.itemText(i) for i in range(self.count())])

    def save_index(self):
        if self.count():
            self.settings.setValue(
                "curr_index_" + self.settings_name,
                self.currentIndex())

    def save_current_text(self):
        if self.count():
            self.settings.setValue(
                "name" + self.settings_name,
                self.currentText())
# -----------------------------------------------------------------------------


class CustomTableWidget(QtWidgets.QTableWidget):
    def __init__(self, parent=None):
        super(CustomTableWidget, self).__init__(parent)
        self.setColumnCount(3)
        self.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        # self.table_widget.setRowHeight(0, 0) 
        self.setHorizontalHeaderLabels(
            ["F, Hz", "A, \u00b0/s", "T, s"])
        self.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.verticalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.total_time = 0

    def set_table(self, file):
        self.setRowCount(0)
        self.total_time = 0
        for line in file:
            f_a_t = list(filter(None, re.split("F|A|T|\n", line)))
            self.add_and_fill_row(f_a_t)

    def add_and_fill_row(self, f_a_t: list):
        if (len(f_a_t) == 3 and all([item.isdecimal()
                                     for item in f_a_t])):
            self.setRowCount(self.rowCount() + 1)
            for j in range(3):
                item = QtWidgets.QTableWidgetItem(f_a_t[j])
                item.setTextAlignment(
                    QtCore.Qt.AlignmentFlag.AlignCenter)
                self.setItem(self.rowCount() - 1, j, item)
            self.total_time += int(f_a_t[-1])

    # get_current_F = lambda self: int(self.item(self.currentRow(), 0).data(QtCore.Qt.ItemDataRole.EditRole))
    def get_current_F(self):
        return int(self.item(
            self.currentRow(), 0).data(
                QtCore.Qt.ItemDataRole.EditRole))

    def get_current_A(self):
        return int(self.item(
            self.currentRow(), 1).data(
                QtCore.Qt.ItemDataRole.EditRole))

    def get_current_T(self):
        return int(self.item(
            self.currentRow(), 2).data(
                QtCore.Qt.ItemDataRole.EditRole)) * 1000
# ----------------------------------------------------------------------------
#
# ----------------------------------------------------------------------------
#
# ----------------------------------------------------------------------------


if __name__ == "__main__":
    import PyQt_ApplicationClass
    app = QtWidgets.QApplication(sys.argv)
    window = PyQt_ApplicationClass.AppWindow()
    sys.exit(app.exec())


#    def plot_fft_median(self, freq_data: np.ndarray, special_points: np.ndarray, folder: str):
#         self.logger.info("Final median plot")
#         # self.append_fft_plot_tab()
#         self.last_tab_layout = QtWidgets.QGridLayout(spacing=0)
#         self.last_tab_layout.addWidget(self.groupbox, 0, 1, 2, 1) 

#         index = self.count() - 1
#         self.tab_widget_page_list.append(QtWidgets.QWidget(self))
#         self.append_amp_plot()
#         self.last_tab_layout.addWidget(self.amp_plot_list[index], 0, 0, 1, 1)
#         self.append_phase_plot()
#         self.last_tab_layout.addWidget(self.phase_plot_list[index], 1, 0, 1, 1)
#         self.tab_widget_page_list[-1].setLayout(self.last_tab_layout)
#         self.addTab(
#             self.tab_widget_page_list[-1], f"ЧХ &{index + 1}")  # FC 
#         self.setTabText(self.count() - 1, "&АФЧХ (средний)")  # FC average
#         self.setCurrentIndex(self.count() - 1)
  
#         # self.del_btn.clicked.connect(
#         #     self.text_())
#         self.name_info.setText(folder)

#         self.x = freq_data[:, -4]  # [1, 5, 13, 20, 55, 62, 80]
#         # # amp =np.array([96, 97, 108, 112, 113, 114, 121, 127, 136, 163, 166, 191, 219, 240, 258, 284, 400, 371, 450, 699, 791, 1154, 631, 697, 722, 743, 834, 823, 918, 995, 1022, 1125, 1220, 1244, 1373, 1468, 1618, 1851, 1865, 2166, 2548, 2539, 3409, 3521, 3514, 4220, 3473, 2573, 3081, 3028, 3028, 3230, 3056, 3132, 3102, 3621, 3669, 4561])/100
#         self.y = freq_data[:, -3]  # np.array([1, 1.2, 1.3, 1.5, 5, .5, .35])
#         # self.amp_curves[-1].setData(self.x, self.y)
#         # self.plot_2d_scatter(self.time_plot, self.x, self.y)
#         # self.cursor = Qt.BlankCursor
#         # self.amp_plot_list[-1].setCursor(self.cursor)  # self.cursor ##########################

#         # Add lines
#         # self.amp_plot_list[-1].addItem(self.infinite_line_x, ignoreBounds=True)
#         # self.amp_plot_list[-1].addItem(self.infinite_line_y, ignoreBounds=True)
    
#         # self.amp_plot_list[-1].addItem(self.cursorlabel)#################################
#         # self.proxy = pg.SignalProxy(
#         #     self.amp_plot_list[-1].scene().sigMouseMoved, delay=0.15,##############################
#         #     rateLimit=15, slot=self.update_crosshair)

#         # self.proxy = pg.SignalProxy(
#             # self.phase_plot_list[-1].scene().sigMouseMoved, delay=0.15,
#             # rateLimit=15, slot=self.update_crosshair) ####################
#         self.mouse_x = None
#         self.mouse_y = None
#         # # self.plot_fft(True)
#         for i in range(self.GYRO_NUMBER):
#             self.amp_curves[-1 - i].setData(freq_data[:, -4],
#                                         freq_data[:, -3])
#             self.phase_curves[-1 - i].setData(freq_data[:, -4],
#                                         freq_data[:, -2])
#         self.amp_plot_list[-1].autoRange()
#     # def plot_2d_scatter(self, plot, x, y, color=(66, 245, 72)):
#     #     brush = pg.mkBrush(color)
#     #     scatter = pg.ScatterPlotItem(size=5, brush=brush)
#     #     scatter.addPoints(x, y)
#     #     plot.addItem(scatter)